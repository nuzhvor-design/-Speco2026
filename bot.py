import os
import re
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.comments import Comment
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# ========== НАСТРОЙКИ ==========
TOKEN = os.environ.get("TELEGRAM_TOKEN")
if not TOKEN:
    raise ValueError("Переменная окружения TELEGRAM_TOKEN не установлена!")

EXCEL_FILE = "SPECO 2026.xlsx"

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)
# ================================

# ----- Категории расходов (ключевое слово -> столбец) -----
EXPENSE_CATEGORY_MAP = {
    "продукты": "B",
    "едавнедома": "C", "кафе": "C", "ресторан": "C",
    "ярослав": "D",
    "транспорт": "E", "общественный": "E",
    "мояработа": "F", "работа": "F",
    "платежи": "G", "кредит": "G", "касарина": "G", "яндекс": "G",
    "подписки": "G", "подписка": "G", "коммуналка": "G",
    "квартплата": "G", "телефон": "G", "налог": "G",
    "гардероб": "H", "одежда": "H",
    "крупныепокупки": "I", "техника": "I",
    "хозтовары": "J", "бытоваяхимия": "J",
    "косметика": "K",
    "салоны": "L", "парикмахерская": "L",
    "здоровье": "M", "лекарства": "M",
    "подарки": "N",
    "психотерапевт": "O", "психолог": "O",
    "спорт": "P", "фитнес": "P",
    "магия": "Q", "благотворительность": "Q", "блг": "Q",
    "развлечения": "R", "кино": "R",
    "путешествия": "S", "отпуск": "S",
}

# ----- Категории доходов для листа "ИТОГО" -----
INCOME_CATEGORY_MAP = {
    "зарплата":   ("AG", "Я"),
    "аванс":      ("AG", "Я"),
    "шабашка":    ("AG", "Я"),
    "алименты":   ("AH", "Алименты"),
}

MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def parse_date_from_string(date_str: str) -> datetime | None:
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None

def get_month_sheet_name(date: datetime) -> str:
    return f"{date.month:02d}"

def find_row_for_date(worksheet, target_date: datetime) -> int | None:
    target_day = target_date.day
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=1)
        if isinstance(cell.value, datetime) and cell.value.day == target_day:
            return row
        if isinstance(cell.value, (int, float)):
            try:
                from openpyxl.utils.datetime import from_excel
                dt = from_excel(cell.value)
                if dt.day == target_day:
                    return row
            except:
                pass
    return None

def find_month_row_on_itto(worksheet, month_name: str) -> int | None:
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=1)
        if cell.value and str(cell.value).strip() == month_name:
            return row
    return None

def add_value_to_cell_with_comment(sheet, col_letter: str, row_num: int, amount: float, comment_text: str = None):
    """
    Прибавляет сумму к ячейке и, если comment_text задан, добавляет комментарий.
    comment_text может быть пустой строкой или None – тогда комментарий не создаётся.
    """
    cell = sheet[f"{col_letter}{row_num}"]
    if cell.data_type == 'f':
        return False, f"Ячейка {col_letter}{row_num} содержит формулу. Бот не может её изменить."
    old = cell.value if cell.value is not None else 0
    try:
        new = float(old) + amount
    except (TypeError, ValueError):
        return False, f"В ячейке {col_letter}{row_num} не число: {old}"
    cell.value = new

    # Добавляем комментарий только если передан непустой текст
    if comment_text and comment_text.strip():
        new_comment = comment_text.strip()
        if cell.comment:
            new_comment = f"{cell.comment.text}\n{new_comment}"
        cell.comment = Comment(new_comment, "Telegram бот")

    return True, f"✅ Добавлено {amount:.2f} руб. в {col_letter}{row_num}"

def update_income_summary(amount: float, keyword: str, comment: str, month_date: datetime):
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        return False, f"Файл {EXCEL_FILE} не найден на сервере."
    except Exception as e:
        return False, f"Ошибка открытия файла: {e}"

    if "ИТОГО" not in wb.sheetnames:
        wb.close()
        return False, "Лист ИТОГО не найден."
    ws_itto = wb["ИТОГО"]
    month_name = MONTHS_RU[month_date.month - 1]
    row_num = find_month_row_on_itto(ws_itto, month_name)
    if row_num is None:
        wb.close()
        return False, f"Не найдена строка для месяца {month_name} на листе ИТОГО."
    if keyword in INCOME_CATEGORY_MAP:
        col, _ = INCOME_CATEGORY_MAP[keyword]
    else:
        col = "AI"
    # Для дохода комментарий передаём только если он есть
    success, msg = add_value_to_cell_with_comment(ws_itto, col, row_num, amount, comment)
    if success:
        wb.save(EXCEL_FILE)
    wb.close()
    return success, msg

def process_operation(line: str, target_date: datetime) -> str:
    line = line.strip()
    if not line:
        return ""

    # ---- Доход ----
    income_match = re.match(r'^\+(\d+(?:[.,]\d+)?)\s+&(.+)$', line, re.IGNORECASE)
    if not income_match:
        income_match = re.match(r'^(?:доход|income)\s+(\d+(?:[.,]\d+)?)\s+&(.+)$', line, re.IGNORECASE)
    if income_match:
        amount = float(income_match.group(1).replace(',', '.'))
        keyword_and_comment = income_match.group(2).strip()
        # Разделяем на ключевое слово и остаток (комментарий)
        parts = keyword_and_comment.split(maxsplit=1)
        keyword = parts[0].lower() if parts else ""
        comment = parts[1] if len(parts) > 1 else ""
        month_sheet = get_month_sheet_name(target_date)

        try:
            wb = load_workbook(EXCEL_FILE)
        except FileNotFoundError:
            return f"❌ Файл {EXCEL_FILE} не найден на сервере."
        except Exception as e:
            return f"❌ Ошибка открытия файла: {e}"

        if month_sheet not in wb.sheetnames:
            wb.close()
            return f"❌ Лист {month_sheet} не найден для даты {target_date.strftime('%d.%m.%Y')}"
        ws_month = wb[month_sheet]
        row = find_row_for_date(ws_month, target_date)
        if row is None:
            wb.close()
            return f"❌ Не найдена строка для {target_date.day}-го дня на листе {month_sheet}"

        # Записываем доход в столбец U (комментарий опционально)
        success, msg_month = add_value_to_cell_with_comment(ws_month, "U", row, amount, comment)
        if not success:
            wb.close()
            return msg_month
        wb.save(EXCEL_FILE)
        wb.close()

        # Обновляем ИТОГО (передаём тот же комментарий)
        success_itto, msg_itto = update_income_summary(amount, keyword, comment, target_date)
        if success_itto:
            return f"✅ Доход:\n   {msg_month}\n   {msg_itto}"
        else:
            return f"⚠️ Доход в лист месяца записан, но ошибка ИТОГО:\n   {msg_itto}"

    # ---- Расход ----
    match = re.match(r'^(\d+(?:[.,]\d+)?)\s+&([a-zA-Zа-яА-ЯёЁ]+)(?:\s+(.*))?$', line, re.IGNORECASE)
    if match:
        amount = float(match.group(1).replace(',', '.'))
        keyword = match.group(2).lower()
        comment = match.group(3) if match.group(3) else ""  # может быть пустым
        if keyword not in EXPENSE_CATEGORY_MAP:
            return f"❌ Неизвестная категория расхода: {keyword}"
        col = EXPENSE_CATEGORY_MAP[keyword]
        month_sheet = get_month_sheet_name(target_date)

        try:
            wb = load_workbook(EXCEL_FILE)
        except FileNotFoundError:
            return f"❌ Файл {EXCEL_FILE} не найден на сервере."
        except Exception as e:
            return f"❌ Ошибка открытия файла: {e}"

        if month_sheet not in wb.sheetnames:
            wb.close()
            return f"❌ Лист {month_sheet} не найден для даты {target_date.strftime('%d.%m.%Y')}"
        ws = wb[month_sheet]
        row = find_row_for_date(ws, target_date)
        if row is None:
            wb.close()
            return f"❌ Не найдена строка для {target_date.day}-го дня на листе {month_sheet}"

        # Передаём комментарий (может быть пустым)
        success, msg = add_value_to_cell_with_comment(ws, col, row, amount, comment)
        if success:
            wb.save(EXCEL_FILE)
        wb.close()
        return msg

    return f"❌ Неверный формат строки: {line}"

# ========== ОБРАБОТЧИКИ TELEGRAM ==========
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "💰 Бот для учёта расходов и доходов\n\n"
        "📆 **Формат сообщения**:\n"
        "Первая строка — дата ДД.ММ.ГГГГ (необязательно, если пропустить – используется сегодня)\n"
        "Затем каждая новая строка — операция:\n"
        "  • Расход: `сумма &категория [комментарий]`\n"
        "  • Доход: `+сумма &ключ [комментарий]`\n\n"
        "Пример:\n"
        "10.07.2026\n"
        "500 &продукты хлеб\n"
        "+30000 &зарплата аванс\n"
        "120 &транспорт\n\n"
        "Комментарий не обязателен – если его нет, ячейка остаётся без примечания.\n\n"
        "Доступные категории расходов:\n"
        "продукты, едавнедома, ярослав, транспорт, работа, платежи,\n"
        "гардероб, крупныепокупки, хозтовары, косметика, салоны,\n"
        "здоровье, подарки, психотерапевт, спорт, магия,\n"
        "развлечения, путешествия\n\n"
        "Ключевые слова для доходов:\n"
        "зарплата, аванс, шабашка → столбец 'Я' (AG)\n"
        "алименты → 'Алименты' (AH)\n"
        "остальное → 'Подарки' (AI)"
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        return
    lines = text.splitlines()
    if not lines:
        return

    # Определяем дату
    first_line = lines[0].strip()
    target_date = parse_date_from_string(first_line)
    start_index = 0
    if target_date:
        start_index = 1
    else:
        target_date = datetime.now()
        start_index = 0

    results = []
    for idx in range(start_index, len(lines)):
        line = lines[idx].strip()
        if not line:
            continue
        res = process_operation(line, target_date)
        if res:
            results.append(res)

    if results:
        await update.message.reply_text("\n\n".join(results))
    else:
        await update.message.reply_text("Не найдено операций для обработки.")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)

# ========== ЗАПУСК БОТА ==========
def main():
    logger.info("Запуск бота...")
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    logger.info("Бот запущен и готов к работе!")
    app.run_polling()

if __name__ == "__main__":
    main()
